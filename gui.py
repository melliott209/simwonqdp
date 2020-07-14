import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import tkcalendar
import openpyxl as xl
import datetime
import random
import data

#============================================================
# Constants
#============================================================

WIDGET_WIDTH = 15
FILE_LABEL_DEFAULT = "None Chosen"
DEFAULT_PADDING = 5
OUT_TOL_MIN = 0.10
OUT_TOL_MAX = 0.40

#============================================================
# Functions
#============================================================

# Reads line from INI file and returns a list
def readList(iniFile):
    string = iniFile.readline()
    startIndex = string.find("=") + 1
    return string[startIndex:].split(",")

# Chooses a file and returns it to the button that asked for it
def chooseFile(label):
    fileName = filedialog.askopenfilename(title="Select File")
    label.insert(0,fileName)
    label.xview(tk.END)
    return

# Chooses a directory and returns it to the button that asked for it
def chooseDir(label):
    dirName = filedialog.askdirectory(title="Choose Directory")
    label.insert(0,dirName)
    return

def partNameCallback(partsList,partNoList,selected,label):
    if selected == '':
        return
    label.config(text=partNoList[partsList.index(selected)])

def dateCallback(julianLabel,date):
    eoy = datetime.date(2019,12,31)
    delta = (date - eoy).days
    julianLabel.config(text=str(delta))

# Loads an excel file from a directory and adds data to the list of parts(samples)
def loadDataPoint(direct, samples):
    if direct != '':
        wb = xl.load_workbook(direct)
        ws = wb.active
        samples.append(data.Part())
        rowlist = list()
        for row in ws.iter_rows(1,ws.max_row):
            for cell in row:
                if cell.value == 'Name':
                    rowlist.append(cell.row)
        for row in rowlist:
            cell = ws.cell(row+1,1)
            while cell.value is not None:
                samples[len(samples)-1].addData(cell.value,
                                                cell.offset(0,1).value,
                                                cell.offset(0,2).value,
                                                cell.offset(0,3).value,
                                                cell.offset(0,4).value[1:],
                                                cell.offset(0,5).value)
                cell = cell.offset(row=1,column=0)

# Loads data from directories given using loadDataPoint()
def loadData(dirList, samples):
    for directory in dirList:
        loadDataPoint(directory, samples)

def writeCSV(header,sample,directory,outfileName):
    with open(directory + '\\' + outfileName + '.csv', 'w') as outFile:
        outFile.write(header)
        for point in sample.getCSV():
            outFile.write(point)

def getFileHeader(partname, partno, date, julian, timecode, supplier, supCode, sampleNo):
    h = int(timecode[:2])
    m = int(timecode[2:])
    time = datetime.datetime(100,1,1,h,m)
    timeoffset = (int(sampleNo[:1])-1) * 15
    time = time + datetime.timedelta(minutes=timeoffset)
    timeString = time.strftime("%I:%M:%S %p")
    returnString = 'Part Number,Part Name,Sample Number\n' + partno + ',' + partname + ',' + sampleNo + '\n' + 'Time,Date\n' + timeString + ',' + date + '\n' + 'Supplier Code,Supplier Name\n' + supCode + ',' + supplier + '\n\n' + 'Feature Label,Feature Type,Nominal X,Actual X,Deviation X,Nominal Y,Actual Y,Deviation Y,Nominal Z,Actual Z,Deviation Z,Diameter/Length,Width\n'
    return returnString

def generateSample(sampleList):
    newPart = data.Part()
    newPart.copyFrom(sampleList[0])
    
    firstSample = sampleList[0].getData()
    for datapoint in firstSample:
        index = firstSample.index(datapoint)

        minimum = datapoint.getX()["Measured"]
        maximum = datapoint.getX()["Measured"]
        for sample in sampleList:
            value = sample.getData()[index].getX()["Measured"]
            if value < minimum:
                minumum = value
            if value > maximum:
                maximum = value
        randValue = round(random.uniform(minimum,maximum),3)
        newPart.getPointByLabel(datapoint.getLabel()).setX("Measured",randValue)

        minimum = datapoint.getY()["Measured"]
        maximum = datapoint.getY()["Measured"]
        for sample in sampleList:
            value = sample.getData()[index].getY()["Measured"]
            if value < minimum:
                minumum = value
            if value > maximum:
                maximum = value
        randValue = round(random.uniform(minimum,maximum),3)
        newPart.getPointByLabel(datapoint.getLabel()).setY("Measured",randValue)

        minimum = datapoint.getZ()["Measured"]
        maximum = datapoint.getZ()["Measured"]
        for sample in sampleList:
            value = sample.getData()[index].getZ()["Measured"]
            if value < minimum:
                minumum = value
            if value > maximum:
                maximum = value
        randValue = round(random.uniform(minimum,maximum),3)
        newPart.getPointByLabel(datapoint.getLabel()).setZ("Measured",randValue)
        
        if datapoint.getType == "Hole":
            minimum = datapoint.getDia()["Measured"]
            maximum = datapoint.getDia()["Measured"]
            for sample in sampleList:
                value = sample.getData()[index].getDia()["Measured"]
                if value < minimum:
                    minimum = value
                if value > maximum:
                    maximum = value
            randValue = round(random.uniform(minimum,maximum),3)
            newPart.getPointByLabel(datapoint.getLabel()).setDia("Measured",randValue)

        elif datapoint.getType == "Slot":
            minimum = datapoint.getLen()["Measured"]
            maximum = datapoint.getLen()["Measured"]
            for sample in sampleList:
                value = sample.getData()[index].getLen()["Measured"]
                if value < minimum:
                    minimum = value
                if value > maximum:
                    maximum = value
            randValue = round(random.uniform(minimum,maximum),3)
            newPart.getPointByLabel(datapoint.getLabel()).setLen("Measured",randValue)

            minimum = datapoint.getWid()["Measured"]
            maximum = datapoint.getWid()["Measured"]
            for sample in sampleList:
                value = sample.getData()[index].getWid()["Measured"]
                if value < minimum:
                    minimum = value
                if value > maximum:
                    maximum = value
            randValue = round(random.uniform(minimum,maximum),3)
            newPart.getPointByLabel(datapoint.getLabel()).setWid("Measured",randValue)

    return newPart

def adjustValues(sample):
    for datapoint in sample.getData():
        nominal = float(datapoint.getX()["Nominal"])
        value = float(datapoint.getX()["Measured"])
        tolerance = float(datapoint.getX()["Tolerance"])
        if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
            newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setX("Measured",newValue)
            datapoint.setX("Deviation",newValue - nominal)
        elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
            newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setX("Measured",newValue)
            datapoint.setX("Deviation",newValue - nominal)

        nominal = float(datapoint.getY()["Nominal"])
        value = float(datapoint.getY()["Measured"])
        tolerance = float(datapoint.getY()["Tolerance"])
        if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
            newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setY("Measured",newValue)
            datapoint.setY("Deviation",newValue - nominal)
        elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
            newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setY("Measured",newValue)
            datapoint.setY("Deviation",newValue - nominal)

        nominal = float(datapoint.getZ()["Nominal"])
        value = float(datapoint.getZ()["Measured"])
        tolerance = float(datapoint.getZ()["Tolerance"])
        if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
            newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setZ("Measured",newValue)
            datapoint.setZ("Deviation",newValue - nominal)
        elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
            newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
            datapoint.setZ("Measured",newValue)
            datapoint.setZ("Deviation",newValue - nominal)

        if datapoint.getType() == "Hole":
            if type(datapoint.getDia()["Measured"]) == str:
                continue
            nominal = float(datapoint.getDia()["Nominal"])
            value = float(datapoint.getDia()["Measured"])
            tolerance = float(datapoint.getDia()["Tolerance"])
            if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
                newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setDia("Measured",newValue)
                datapoint.setDia("Deviation",newValue - nominal)
            elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
                newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setDia("Measured",newValue)
                datapoint.setDia("Deviation",newValue - nominal)

        elif datapoint.getType() == "Slot":
            nominal = float(datapoint.getLen()["Nominal"])
            value = float(datapoint.getLen()["Measured"])
            tolerance = float(datapoint.getLen()["Tolerance"])
            if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
                newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setLen("Measured",newValue)
                datapoint.setLen("Deviation",newValue - nominal)
            elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
                newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setLen("Measured",newValue)
                datapoint.setLen("Deviation",newValue - nominal)

            nominal = float(datapoint.getWid()["Nominal"])
            value = float(datapoint.getWid()["Measured"])
            tolerance = float(datapoint.getWid()["Tolerance"])
            if value-nominal > tolerance + (tolerance * OUT_TOL_MAX):
                newValue = round(nominal + (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setWid("Measured",newValue)
                datapoint.setWid("Deviation",newValue - nominal)
            elif value-nominal < -tolerance - (tolerance * OUT_TOL_MAX):
                newValue = round(nominal - (tolerance + (tolerance * random.uniform(OUT_TOL_MIN,OUT_TOL_MAX))), 3)
                datapoint.setWid("Measured",newValue)
                datapoint.setWid("Deviation",newValue - nominal)

def addHardware(sample):
    for datapoint in sample.getData():
        for label in hardwareParts:
            if datapoint.getLabel() == label:
                datapoint.setDia('Measured',hardware[hardwareParts.index(label)])

# Main command for Generate Samples button on GUI
def processCMM(dirList, numSamples, partname, partno, date, julian, timecode, supplier, supcode, outDir, allGen):

    # Check for errors in the entry fields
    flag = False
    for directory in dirList:
        if directory != '':
            flag = True
    if flag == False:
        tk.messagebox.showerror(title='Error',message='No input files chosen!')
        return
    if outDir == '':
        messagebox.showerror(title='Error',message='No output directory chosen!')
        return

    # TODO: Add more errors for missing entries

    # Create List to hold part "samples" and load it from selected directories
    Samples = list()
    loadData(dirList,Samples)

    # Counter for number of samples generated; initialize to 0
    count = 0

    if allGen == 0:
        while count < int(numSamples):
            if count < len(dirList):
                if dirList[count] != '':
                    sampleNo = str(count+1) + '-' + julian
                    outFileName = partno + '_SampleID_' + sampleNo
                    adjustValues(Samples[count])
                    addHardware(Samples[count])
                    writeCSV(getFileHeader(partname,partno,date,julian,timecode,supplier,supcode,sampleNo),Samples[count],outDir,outFileName)
                    count += 1
                else:
                    part = generateSample(Samples)
                    adjustValues(part)
                    addHardware(part)
                    sampleNo = str(count+1) + '-' + julian
                    outFileName = partno + '_SampleID_' + sampleNo
                    writeCSV(getFileHeader(partname,partno,date,julian,timecode,supplier,supcode,sampleNo),part,outDir,outFileName)
                    count += 1
    else:
        while count < int(numSamples):
            part = generateSample(Samples)
            adjustValues(part)
            addHardware(part)
            sampleNo = str(count+1) + '-' + julian
            outFileName = partno + '_SampleID_' + sampleNo
            writeCSV(getFileHeader(partname,partno,date,julian,timecode,supplier,supcode,sampleNo),part,outDir,outFileName)
            count += 1
        
#============================================================
# Load and Process Input Files
#============================================================

# TODO: In INI file, new list of parts and partnos from Master List with all subs/singles,etcy
# Process INI file into variables
ini_parts = open("parts.ini","r")
parts = readList(ini_parts)
partnos = readList(ini_parts)
endassy = readList(ini_parts)
line = readList(ini_parts)
cause = readList(ini_parts)
rework = readList(ini_parts)
supplier = readList(ini_parts)
process = readList(ini_parts)
shift = readList(ini_parts)
dept = readList(ini_parts)
hardwareParts = readList(ini_parts)
hardware = readList(ini_parts)
ini_parts.close()

#============================================================
# GUI Code
#============================================================

# Create window
window = tk.Tk()
window.title("Simwon QDP")
window.iconphoto(False, tk.PhotoImage(file='SimwonAmerica-Logo2.png'))
window.resizable(False,False)
window.grid_columnconfigure(0,weight=1)

## Create Tab holder
tabCtl = ttk.Notebook(window)
tabCtl.grid_columnconfigure(0,weight=1)

## Create tabs
tabAdd = tk.Frame(tabCtl,padx=DEFAULT_PADDING,pady=DEFAULT_PADDING)
tabEdit = tk.Frame(tabCtl,padx=DEFAULT_PADDING,pady=DEFAULT_PADDING)
tabEdit.grid_columnconfigure(0,weight=1)

#============================================================
# Add Entry Tab
#============================================================

#### Add "General" Container
generalLabel = tk.Label(tabAdd,text="General",relief=tk.GROOVE)
general = tk.Frame(tabAdd,relief=tk.GROOVE,borderwidth=1,padx=DEFAULT_PADDING,pady=DEFAULT_PADDING)

dateLabel = tk.Label(general,text="Date",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=0,column=0)
dateEntry = tkcalendar.DateEntry(general,width=WIDGET_WIDTH).grid(row=0,column=1)
partNameLabel = tk.Label(general,text="Part Name",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=1,column=0)
partNameEntry = ttk.Combobox(general,width=WIDGET_WIDTH,values=parts).grid(row=1,column=1)
partNoLabel = tk.Label(general,text="Part Number",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=2,column=0)
partNoEntry = tk.Label(general,width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=2,column=1,sticky='ew')
endAssyLabel = tk.Label(general,text="End Assembly",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=3,column=0)
endAssyEntry = ttk.Combobox(general,width=WIDGET_WIDTH,values=endassy).grid(row=3,column=1)
lineLabel = tk.Label(general,text="Line",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=4,column=0)
lineEntry = ttk.Combobox(general,width=WIDGET_WIDTH,values=line).grid(row=4,column=1)

generalLabel.grid(row=0,column=0,sticky='ew')
general.grid(row=1,column=0,sticky='n')

#### Add "Defect" Container
defectLabel = tk.Label(tabAdd,text="Defect",relief=tk.GROOVE)
defect = tk.Frame(tabAdd,relief=tk.GROOVE,borderwidth=1,padx=DEFAULT_PADDING,pady=DEFAULT_PADDING)

descLabel = tk.Label(defect,text="Description",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=0,column=0,sticky='n')
descEntry = tk.Text(defect,width=WIDGET_WIDTH,height=5).grid(row=0,column=1)
causeLabel = tk.Label(defect,text="Cause",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=1,column=0)
causeEntry = ttk.Combobox(defect,width=WIDGET_WIDTH,values=cause).grid(row=1,column=1,sticky='ew')
qtyLabel = tk.Label(defect,text="Quantity",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=2,column=0)
qtyEntry = tk.Spinbox(defect,from_=0,to=10000,width=WIDGET_WIDTH).grid(row=2,column=1,sticky='ew')
decisionLabel = tk.Label(defect,text="Rework/Scrap",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=3,column=0)
decisionEntry = ttk.Combobox(defect,width=WIDGET_WIDTH,values=rework).grid(row=3,column=1,sticky='ew')
costLabel = tk.Label(defect,text="Scrap Cost",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=4,column=0)
costEntry = tk.Entry(defect,width=WIDGET_WIDTH).grid(row=4,column=1,sticky='ew')

defectLabel.grid(row=0,column=1,sticky='ew')
defect.grid(row=1,column=1)

# TODO: Add "Images" Container

# Add "Accountability" Container
accountLabel = tk.Label(tabAdd,text="Accountability",relief=tk.GROOVE)
account = tk.Frame(tabAdd,relief=tk.GROOVE,borderwidth=1,padx=DEFAULT_PADDING,pady=DEFAULT_PADDING)

supplierLabel = tk.Label(account,text="Supplier",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=0,column=0)
supplierEntry = ttk.Combobox(account,width=WIDGET_WIDTH,values=supplier).grid(row=0,column=1,sticky='ew')
processLabel = tk.Label(account,text="Process",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=1,column=0)
processEntry = ttk.Combobox(account,width=WIDGET_WIDTH,values=process).grid(row=1,column=1,sticky='ew')
shiftLabel = tk.Label(account,text="Shift",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=2,column=0)
shiftEntry = ttk.Combobox(account,width=WIDGET_WIDTH,values=shift).grid(row=2,column=1,sticky='ew')
deptLabel = tk.Label(account,text="Responsible Dept.",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=3,column=0)
deptEntry = ttk.Combobox(account,width=WIDGET_WIDTH,values=dept).grid(row=3,column=1,sticky='ew')
qcLabel = tk.Label(account,text="QC Inspector",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=4,column=0)
qcEntry = tk.Entry(account,width=WIDGET_WIDTH).grid(row=4,column=1,sticky='ew')
personLabel = tk.Label(account,text="Responsible Person",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=5,column=0)
personEntry = tk.Entry(account,width=WIDGET_WIDTH).grid(row=5,column=1,sticky='ew')
remarkLabel = tk.Label(account,text="Remarks",width=WIDGET_WIDTH,relief=tk.GROOVE).grid(row=6,column=0,sticky='n')
remarkEntry = tk.Text(account,width=WIDGET_WIDTH,height=3).grid(row=6,column=1)

accountLabel.grid(row=2,column=1,sticky='ew')
account.grid(row=3,column=1)

### Finalize "Add Entry" tab
tabCtl.add(tabAdd, text="Record Defect")

#============================================================
# Process CMM Tab
#============================================================

inputLabel = tk.Label(tabEdit,text="Input Files",relief=tk.GROOVE,pady=5)
outputLabel = tk.Label(tabEdit,text="Output Options",relief=tk.GROOVE,pady=5)
inputFrame = tk.Frame(tabEdit,relief=tk.GROOVE,borderwidth=1,pady=5)
inputFrame.grid_columnconfigure(0,weight=1)
inputFrame.grid_columnconfigure(1,weight=1)
outputFrame = tk.Frame(tabEdit,relief=tk.GROOVE,borderwidth=1,pady=5)
outputFrame.grid_columnconfigure(0,weight=1)
outputFrame.grid_columnconfigure(1,weight=1)
fileLabel1 = tk.Entry(inputFrame,relief=tk.GROOVE)
fileLabel2 = tk.Entry(inputFrame,relief=tk.GROOVE)
fileLabel3 = tk.Entry(inputFrame,relief=tk.GROOVE)
saveDirLabel = tk.Entry(outputFrame,relief=tk.GROOVE)
fileBtn1 = tk.Button(inputFrame,text="Choose File...",command=lambda: chooseFile(fileLabel1))    # Have to pass lambda in order to pass args
fileBtn2 = tk.Button(inputFrame,text="Choose File...",command=lambda: chooseFile(fileLabel2))
fileBtn3 = tk.Button(inputFrame,text="Choose File...",command=lambda: chooseFile(fileLabel3))
saveDirBtn = tk.Button(outputFrame,text="Choose Directory...",command=lambda: chooseDir(saveDirLabel))
sampleQtyLabel = tk.Label(outputFrame,text="Sample Quantity",relief=tk.GROOVE,pady=5)
sampleQtyEntry = tk.Spinbox(outputFrame,from_=0,to=100)
partNameLab = tk.Label(outputFrame,text="Part Name",relief=tk.GROOVE,pady=5)
partNameEnt = ttk.Combobox(outputFrame,values=parts,state="readonly")
partNoLab = tk.Label(outputFrame,text="Part Number",relief=tk.GROOVE,pady=5)
partNoEnt = tk.Label(outputFrame,text="--",relief=tk.GROOVE,anchor='w')
dateLab = tk.Label(outputFrame,text="Date",relief=tk.GROOVE,pady=5)
dateEnt = tkcalendar.DateEntry(outputFrame,locale='en_US')
partNameEnt.bind("<<ComboboxSelected>>",lambda event:partNameCallback(parts,partnos,partNameEnt.get(),partNoEnt))
julianDateLabel = tk.Label(outputFrame,text="Julian Date Code",relief=tk.GROOVE,pady=5)
julianDateEntry = tk.Label(outputFrame,text="--",relief=tk.GROOVE,pady=5,anchor='w')
dateEnt.bind("<<DateEntrySelected>>",lambda event:dateCallback(julianDateEntry,dateEnt.get_date()))
timeLabel = tk.Label(outputFrame,text="Time (e.g. 1430 => 2:30 PM)",relief=tk.GROOVE,pady=5)
timeEntry = tk.Entry(outputFrame,relief=tk.GROOVE)
supplierLab = tk.Label(outputFrame,text="Supplier",relief=tk.GROOVE,pady=5)
supplierEnt = tk.Label(outputFrame,text="MS Auto",relief=tk.GROOVE,pady=5,anchor='w')
supCodeLabel = tk.Label(outputFrame,text="Supplier Code",relief=tk.GROOVE,pady=5)
supCodeEntry = tk.Label(outputFrame,text="131307",relief=tk.GROOVE,pady=5,anchor='w')
allGen = tk.IntVar()
allGenEntry = tk.Checkbutton(outputFrame,text="Generated Samples Only",variable=allGen)
genBtn = tk.Button(outputFrame,text="Generate Samples",padx=10,pady=5,command=lambda: processCMM(list([fileLabel1.get(),fileLabel2.get(),fileLabel3.get()]),sampleQtyEntry.get(),partNameEnt.get(),partNoEnt.cget("text"),dateEnt.get(),julianDateEntry.cget("text"),timeEntry.get(),supplierEnt.cget("text"),supCodeEntry.cget("text"),saveDirLabel.get(),allGen))

fileLabel1.grid(row=0,column=1,sticky='nsew')
fileLabel2.grid(row=1,column=1,sticky='nsew')
fileLabel3.grid(row=2,column=1,sticky='nsew')
fileBtn1.grid(row=0,column=0,sticky='ew')
fileBtn2.grid(row=1,column=0,sticky='ew')
fileBtn3.grid(row=2,column=0,sticky='ew')
saveDirBtn.grid(row=0,column=0,sticky='ew')
saveDirLabel.grid(row=0,column=1,sticky='nsew')
sampleQtyLabel.grid(row=1,column=0,sticky='ew')
sampleQtyEntry.grid(row=1,column=1,sticky='nsew')
partNameLab.grid(row=2,column=0,sticky='nsew')
partNameEnt.grid(row=2,column=1,sticky='nsew')
partNoLab.grid(row=3,column=0,sticky='nsew')
partNoEnt.grid(row=3,column=1,sticky='nsew')
dateLab.grid(row=4,column=0,sticky='nsew')
dateEnt.grid(row=4,column=1,sticky='nsew')
julianDateLabel.grid(row=5,column=0,sticky='nsew')
julianDateEntry.grid(row=5,column=1,sticky='nsew')
timeLabel.grid(row=6,column=0,sticky='nsew')
timeEntry.grid(row=6,column=1,sticky='nsew')
supplierLab.grid(row=7,column=0,sticky='nsew')
supplierEnt.grid(row=7,column=1,sticky='nsew')
supCodeLabel.grid(row=8,column=0,sticky='nsew')
supCodeEntry.grid(row=8,column=1,sticky='nsew')
genBtn.grid(row=9,column=0,columnspan=2)
allGenEntry.grid(row=10,column=1,sticky='nsew')

inputLabel.grid(row=0,column=0,sticky='nsew')
outputLabel.grid(row=2,column=0,sticky='nsew')

inputFrame.grid(row=1,column=0,sticky='nsew')
outputFrame.grid(row=3,column=0,sticky="nsew")

### Finalize "Process CMM" tab
tabCtl.add(tabEdit, text="Process CMM")

## Finalize tab holder
tabCtl.grid(row=0,column=0)

#============================================================
# Hand Program over to Window's Main Loop
#============================================================

# Start main loop
window.mainloop()
